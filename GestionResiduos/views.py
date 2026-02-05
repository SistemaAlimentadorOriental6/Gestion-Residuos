from collections import defaultdict
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout
from GestionResiduos.models import FormularioPerfil1, FormularioPerfil2, GrupoResiduo, AutorizacionSalida, ResiduoPrecio
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden
from django.utils.safestring import mark_safe
from django.core.serializers.json import DjangoJSONEncoder
import json
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from django.conf import settings
import os
from openpyxl.utils import get_column_letter
from django.db.models.functions import ExtractMonth, ExtractYear
from decimal import Decimal
import uuid
from django.core.paginator import Paginator
from django.utils import timezone



def inicio(request):
    return render(request, 'base.html')







@login_required
def formularioResiduos(request):
    es_sergio = request.user.username == '1036619811'

    if request.method == 'POST':
        seleccionados = request.POST.getlist('seleccionados')
        proveedor = request.POST.get("proveedor")
        if not seleccionados:
            messages.warning(request, "No seleccionaste ningún residuo.")
            return redirect('formularioResiduos')

        registro_exitoso = False  # Bandera para controlar si al menos un residuo fue registrado

        for clave in seleccionados:
            print(f"Caalve recibida{clave}")
            try:
                tipo_residuo, residuo = clave.split(' - ', 1)
            except ValueError:
                continue  # ignorar claves mal formateadas

            peso_raw = request.POST.get(f'peso_{clave}', '0').replace(',', '.')
            cantidad_raw = request.POST.get(f'cantidad_{clave}', '0')

            try:
                peso = Decimal(peso_raw)
            except:
                peso = Decimal('0.0')

            try:
                cantidad = Decimal(cantidad_raw)
            except:
                cantidad = Decimal('0.0')

            hoy = timezone.localdate()

            posibles_grupos = GrupoResiduo.objects.filter(
                completado=False,
                creado_en__date=hoy
            )
            grupo = None

            for posible in posibles_grupos:
                existe_1 = FormularioPerfil1.objects.filter(grupo_codigo=posible).exists()
                existe_2 = FormularioPerfil2.objects.filter(grupo_codigo=posible).exists()

                if es_sergio and not existe_1:
                    perfil2 = FormularioPerfil2.objects.filter(
                        grupo_codigo=posible, tipo_residuo=tipo_residuo, residuo=residuo
                    ).first()
                    if perfil2:  # Eliminamos la validación de tolerancia
                        grupo = posible
                        break

                elif not es_sergio and not existe_2:
                    perfil1 = FormularioPerfil1.objects.filter(
                        grupo_codigo=posible, tipo_residuo=tipo_residuo, residuo=residuo
                    ).first()
                    if perfil1:  # Eliminamos la validación de tolerancia
                        grupo = posible
                        break

            if not grupo:
                if es_sergio:
                    messages.error(request, f"No se encontró grupo compatible para: {residuo}.")
                    continue  # seguimos con el siguiente residuo
                else:
                    # Generar código único para nuevo grupo
                    codigo = f"GRP-{uuid.uuid4()}"
                    grupo = GrupoResiduo.objects.create(codigo=codigo, creado_por=request.user)

            if es_sergio:
                costo_unitario_raw = request.POST.get(f'costo_unitario_{clave}', '0').replace('.', '').replace(',', '')
                costo_total_raw = request.POST.get(f'costo_total_{clave}', '0').replace('.', '').replace(',', '')

                try:
                    costo_unitario = int(costo_unitario_raw or '0')
                except:
                    costo_unitario = 0

                try:
                    costo_total = int(costo_total_raw or '0')
                except:
                    costo_total = 0

                FormularioPerfil1.objects.create(
                    tipo_residuo=tipo_residuo,
                    residuo=residuo,
                    peso=peso,
                    cantidad=int(cantidad),
                    costo_unitario=costo_unitario,
                    costo_total=costo_total,
                    grupo_codigo=grupo,
                    usuario=request.user,
                    proveedor=proveedor
                )

                if not ResiduoPrecio.objects.filter(residuo=residuo).exists():
                    ResiduoPrecio.objects.create(
                        tipo_residuo=tipo_residuo,
                        residuo=residuo,
                        costo_unitario=costo_unitario
                    )

            else:
                FormularioPerfil2.objects.create(
                    tipo_residuo=tipo_residuo,
                    residuo=residuo,
                    peso=peso,
                    cantidad=int(cantidad),
                    grupo_codigo=grupo,
                    usuario=request.user,
                    proveedor=proveedor

                )

            perfil1_ok = FormularioPerfil1.objects.filter(grupo_codigo=grupo).exists()
            perfil2_ok = FormularioPerfil2.objects.filter(grupo_codigo=grupo).exists()
            if perfil1_ok and perfil2_ok:
                grupo.completado = True
                grupo.save()

            registro_exitoso = True  # Al menos un registro se hizo correctamente

        if registro_exitoso:
            messages.success(request, "Residuos registrados correctamente.")
            if es_sergio:
                return redirect('registrosSgi')
            else:
                return redirect('registrosVigilantes')

        return redirect('formularioResiduos')

    # --- GET ---
    residuos_precios = ResiduoPrecio.objects.all()
    precios_dict = {r.residuo: r.costo_unitario for r in residuos_precios}
    precios_json = mark_safe(json.dumps(precios_dict, cls=DjangoJSONEncoder))

    residuos_por_tipo = {
        "Aprovechable": [
            "Acero Inoxidable - KG",
            "Aluminio Limpio - KG",
            "Aluminio Sucio - KG",
            "Bidón Metálico * 55 gal - UND",
            "Bidón Metálico * 55 gal Malo - UND",
            "Bidón Plástico * 5 gal - UND",
            "Bronce - KG",
            "Cartón - KG",
            "Chatarra Metálica - KG",
            "Cobre Contaminado - KG",
            "Cobre Encauchetado - KG",
            "Cobre Rojo - KG",
            "Filtro de Aceite - KG",
            "Hierro Gris - KG",
            "Intercooler Aluminio - KG",
            "Isotanque Base Madera - UND",
            "Isotanque Base Metálica - UND",
            "Inversor 3 KP - UND",
            "Inversor Cargador - KG",
            "Material Mixto - KG",
            "Material para Catalizador - KG",
            "Motor Eléctrico Usado (Bus Articulado) - KG",
            "Panel Solar 1650 * 992 - UND",
            "Papel Archivo - KG",
            "Papel Archivo AZ - KG",
            "Papel Kraft - KG",
            "Papel Mixto - KG",
            "Papel Periódico - KG",
            "Pasta de Archivo AZ - KG",
            "Pasta por Seleccionar - KG",
            "Pasta Seleccionada - KG",
            "Pasta Sillas - KG",
            "Pet por Seleccionar - KG",
            "Pet Transparente - KG",
            "Plástico por Seleccionar - KG",
            "Plástico Seleccionado - KG",
            "Plegadiza Limpia - KG",
            "Plomo - KG",
            "PVC - KG",
            "Radiador Aluminio - KG",
            "Radiador Cobre - KG",
            "Tetra Pak - KG",
            "Vidrio - KG"
        ],

        "Especial": [
            "Bloques de Freno - m³",
            "Fibra de Vidrio - m³",
            "Filtro de Aire - m³",
            "Madera - m³",
            "Residuo Vegetal - m³",
            "Vidrio Panorámico - m³"
        ],    
        
        "Respel": [
            "Agua Contaminada con Hidrocarburo - KG",
            "Filtro de Aceite Cartón - KG",
            "Filtros de Aceite - KG",
            "Grasa Usada - KG",
            "Lodos - KG",
            "Líquido Inflamable (Gasolina) - KG",
            "Luminarias - KG",
            "Material Cortopunzante - KG",
            "Raee - KG",
            "Refrigerante Usado - KG",
            "Residuos Químicos Líquidos - KG",
            "Sólidos Contaminados con Hidrocarburos - KG",
            "Sólidos Contaminados con Pintura - KG",
            "Thinner Contaminado - KG",
            "Thonner - KG"
        ],        
        
        "Respel Aprovechable": [
            "Aceite Usado * gal - UND"
        ],
        
        "Post Consumo": [
            "Batería 27 H - UND",
            "Batería 34 D - UND",
            "Batería 34 H - UND",
            "Batería 34 RST - UND",
            "Batería 4 DLT 4D - UND",
            "Batería Ácido Plomo 24R - UND",
            "Batería Ácido Plomo 30-31H - UND",
            "Batería Ácido Plomo 65-G4-27 - UND",
            "Batería Caja 42 - UND",
            "Batería Grupo 0,7 - UND",
            "Batería Grupo 1 - UND",
            "Batería Grupo 8 - UND",
            "Batería Moto - UND",
            "Batería UPS - UND"
        ]   
    }

    grupos_pendientes = []
    if es_sergio:
        grupos_incompletos = GrupoResiduo.objects.filter(
            completado=False,
            formularioperfil1__isnull=True,
            formularioperfil2__isnull=False
        ).distinct()

        for grupo in grupos_incompletos:
            registro_vigilancia = FormularioPerfil2.objects.filter(grupo_codigo=grupo).first()
            if registro_vigilancia:
                grupos_pendientes.append({
                    'codigo': grupo.codigo,
                    'fecha': registro_vigilancia.fecha,
                    'tipo': registro_vigilancia.tipo_residuo,
                    'residuo': registro_vigilancia.residuo,
                })

    return render(request, 'Residuos/formulario.html', {
        'es_sergio': es_sergio,
        'grupo_codigo': '',
        'precios_residuos_json': precios_json,
        'precios_residuos': precios_dict,  
        'grupos_pendientes': grupos_pendientes,
        'residuos_por_tipo': residuos_por_tipo
    })




@login_required
def listadoAutorizaciones(request):
    if request.user.username != '1112629169':
        return HttpResponseForbidden("No tienes permiso para ver esta página.")

    grupos = GrupoResiduo.objects.filter(completado=True).exclude(autorizacionsalida__isnull=False)
    grupos_alerta = []

    baterias_keywords = ['bateria', 'batería', 'und']

    for grupo in grupos:
        registros_p1 = grupo.formularioperfil1_set.all()
        registros_p2 = grupo.formularioperfil2_set.all()

        residuos_data = defaultdict(lambda: {'perfil1': 0, 'perfil2': 0, 'modo': 'peso', 'proveedor1': None, 'proveedor2': None})

        for r in registros_p1:
            key = r.residuo.lower().strip()
            modo = 'cantidad' if any(k in key for k in baterias_keywords) else 'peso'
            valor_p1 = r.cantidad if modo == 'cantidad' else float(r.peso or 0)
            residuos_data[key]['perfil1'] += valor_p1
            residuos_data[key]['modo'] = modo
            residuos_data[key]['proveedor1'] = r.proveedor  # Asumiendo que "proveedor" es un campo en el modelo

        for r in registros_p2:
            key = r.residuo.lower().strip()
            modo = 'cantidad' if any(k in key for k in baterias_keywords) else 'peso'
            valor_p2 = r.cantidad if modo == 'cantidad' else float(r.peso or 0)
            residuos_data[key]['perfil2'] += valor_p2
            residuos_data[key]['modo'] = modo
            residuos_data[key]['proveedor2'] = r.proveedor  # Asumiendo que "proveedor" es un campo en el modelo

        diferencias = []    
        for residuo, datos in residuos_data.items():
            v1, v2 = datos['perfil1'], datos['perfil2']
            modo = datos['modo']
            proveedor1, proveedor2 = datos['proveedor1'], datos['proveedor2']

            # Comprobar si el proveedor es diferente
            if proveedor1 != proveedor2:
                diferencias.append({
                    'residuo': residuo,
                    'modo': 'Proveedor',
                    'proveedor1': proveedor1,
                    'proveedor2': proveedor2,
                    'mensaje': 'Proveedores diferentes'
                })

            if modo == 'cantidad':
                if v1 != v2:
                    diferencias.append({
                        'residuo': residuo,
                        'modo': 'Cantidad',
                        'valor_p1': v1,
                        'valor_p2': v2,
                        'mensaje': v1 - v2
                    })
            else:  # modo == peso
                if v1 or v2:  # Si hay algún valor
                    referencia = v1 if v1 else 1  # Evitar división por cero
                    dif_pct = abs(v1 - v2) / referencia * 100
                    if dif_pct > 0:  # Solo agregar si hay diferencia real
                        diferencias.append({
                            'residuo': residuo,
                            'modo': 'Peso',
                            'valor_p1': v1,
                            'valor_p2': v2,
                            'diferencia_pct': round(dif_pct, 2)
                        })

        grupos_alerta.append({
            'grupo': grupo,
            'diferencias': diferencias,
            'alerta': bool(diferencias)
        })

    if request.method == 'POST':
        grupo_id = request.POST.get('grupo_id')
        estado = request.POST.get('estado')
        observacion = request.POST.get('observacion', '')
        alerta_peso = request.POST.get('alerta_peso', '')
        print(f"Alerta peso recibida: {alerta_peso}")
        
        diferencia = alerta_peso or request.POST.get('alerta_cant', '')
        print(f"La novedad es: {diferencia}")

        grupo = get_object_or_404(GrupoResiduo, id=grupo_id)

        AutorizacionSalida.objects.create(
            grupo_codigo=grupo,
            estado=estado,
            observacion=observacion or "Ninguna",
            autorizador=request.user,
            diferencia=diferencia or "Ninguna",
        )

        messages.success(request, f'Grupo {grupo.codigo} autorizado correctamente.')
        return redirect('listadoAutorizaciones')

    return render(request, 'residuos/listadoAutorizaciones.html', {
        'grupos_alerta': grupos_alerta
    })





@login_required
def registrosVigilantes(request):
    registros = FormularioPerfil2.objects.filter(usuario=request.user).order_by('-fecha', '-hora')
    return render(request, 'Residuos/registrosVigilantes.html', {'registros': registros})



@login_required
def registrosSgi(request):
    registros = FormularioPerfil1.objects.filter(usuario=request.user).order_by('-fecha', '-hora')
    return render(request, 'Residuos/registrosSergio.html', {'registros': registros})




def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            if user.username == '1112629169':  
                return redirect('listadoAutorizaciones')
            else:  
                return redirect('formularioResiduos')

        else:
            messages.error(request, 'Credenciales inválidas')
    return render(request, 'Login/login.html')




def logout_view(request):
    logout(request)
    return redirect('login')



@login_required
def agregarResiduoPrecio(request):
    if request.method == 'POST':
        tipo_residuo = request.POST['tipo_residuo']
        residuo = request.POST['residuo']
        costo_unitario = int(request.POST.get('costo_unitario', '0').replace('.', '').replace(',', ''))
        
        residuo_costo = ResiduoPrecio.objects.create(
            tipo_residuo = tipo_residuo,
            residuo = residuo,
            costo_unitario = costo_unitario,
        )
        
        residuo_costo.save()
        
        messages.success(request, "Residuo y precio guardado exitosamente")
        return redirect('agregarResiduoPrecio')
    
    return render(request, "Residuos/agregarResiduo.html")




@login_required
def listadoResiduosPrecios(request):
    residuos_precios = ResiduoPrecio.objects.all()
    return render(request, "Residuos/listadoResiduos.html", {'residuos_precios' : residuos_precios})




@login_required
def actualizarResiduoPrecio(request, id):
    residuo_obj = get_object_or_404(ResiduoPrecio, id=id)
    if request.method == 'POST':
        tipo_residuo = request.POST['tipo_residuo']
        residuo = request.POST['residuo']
        costo_unitario = int(request.POST.get('costo_unitario', '0').replace('.', '').replace(',', ''))
        
        residuo_obj.tipo_residuo = tipo_residuo
        residuo_obj.residuo = residuo
        residuo_obj.costo_unitario = costo_unitario
        residuo_obj.save()
        
        residuo_obj.save()
        
        messages.success(request, "Residuo y precio guardado exitosamente")
        return redirect('listadoResiduosPrecio')
    
    return render(request, 'Residuos/actualizarResiduo.html', {
        'residuo': residuo_obj
    })




@login_required
def actualizarCostoTotal(request, registro_id):
    registro = get_object_or_404(FormularioPerfil1, id=registro_id)

    if request.method == 'POST':
        try:
            registro.proveedor = request.POST.get('proveedor', registro.proveedor)
            registro.tipo_residuo = request.POST.get('tipo_residuo', registro.tipo_residuo)
            registro.residuo = request.POST.get('residuo', registro.residuo)

            # Sanitizar numéricos
            def limpiar_num(valor):
                return float(valor.replace('.', '').replace(',', '.')) if valor else 0

            registro.costo_unitario = int(limpiar_num(request.POST.get('costo_unitario')))
            registro.cantidad = int(request.POST.get('cantidad') or 0)
            registro.peso = limpiar_num(request.POST.get('peso'))
            registro.costo_total = int(limpiar_num(request.POST.get('costo_total')))

            registro.save()
            messages.success(request, "Registro actualizado correctamente.")

        except Exception as e:
            messages.error(request, f"Error al actualizar: {str(e)}")

    return redirect('registrosSgi')




@login_required
def actualizarRegistroVigilante(request, registro_id):
    registro = get_object_or_404(FormularioPerfil2, id=registro_id)
    
    if request.method == "POST":
        try:
            registro.proveedor = request.POST.get('proveedor', registro.proveedor)
            registro.tipo_residuo = request.POST.get('tipo_residuo', registro.tipo_residuo)
            registro.residuo = request.POST.get('residuo', registro.residuo)

            def limpiar_num(valor):
                return float(valor.replace('.', '').replace(',', '.')) if valor else 0

            cantidad_raw = request.POST.get('cantidad', '0')
            registro.cantidad = int(float(cantidad_raw.replace(',', '.')))  # <-- Aquí el fix

            registro.peso = limpiar_num(request.POST.get('peso'))
            
            registro.save()
            messages.success(request, "Registro actualizado correctamente.")
            
        except Exception as e:
            messages.error(request, f"Error al actualizar: {str(e)}")
            
        return redirect('registrosVigilantes')


    residuos_por_tipo = {
    "Aprovechable": [...],
    "Especial": [...],
    "Respel": [...],
    "Respel Aprovechable": [...]
    }

    
    return render(request, 'Residuos/registrosVigilantes.html', {
        "residuos_por_tipo": residuos_por_tipo
    })




@login_required
def generarExcel(request, registro_id):
    registro_base = FormularioPerfil1.objects.get(pk=registro_id)
    fecha_base = registro_base.fecha

    mes = fecha_base.month
    año = fecha_base.year
    meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    nombre_archivo = f'Consolidado PMIRS {meses[mes]} {año}.xlsx'
    file_path = os.path.join(settings.MEDIA_ROOT, nombre_archivo)

    registros = FormularioPerfil1.objects.annotate(
        mes=ExtractMonth('fecha'),
        anio=ExtractYear('fecha')
    ).filter(mes=mes, anio=año)

    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidados"
        encabezados = ["Tipo_Residuo", "Residuo", "Fecha", "Peso", "Cantidad", "Costo_Unitario", "Costo_Total"]

        estilo = {
            "fill": PatternFill(start_color="6BA43A", end_color="6BA43A", fill_type="solid"),
            "font": Font(color="000000", bold=True),
            "alignment": Alignment(horizontal="center")
        }

        for col_num, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num, value=encabezado)
            cell.fill = estilo["fill"]
            cell.font = estilo["font"]
            cell.alignment = estilo["alignment"]

        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active
    registros_guardados = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:  # residuo y fecha
            registros_guardados.add((row[1], row[2]))  # clave sin hora

    fila_excel = ws.max_row + 1
    for reg in registros:
        fecha_str = reg.fecha.strftime('%d-%m-%y')
        clave = (reg.residuo, fecha_str)

        if clave in registros_guardados:
            continue

        fila = [
            reg.tipo_residuo,
            reg.residuo,
            fecha_str,
            float(reg.peso),
            float(reg.cantidad),
            float(reg.costo_unitario),
            float(reg.costo_total)
        ]

        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=fila_excel, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal="center")
            if col_idx in [6, 7]:  # Costo_Unitario y Costo_Total
                cell.number_format = '#,##0'    

        fila_excel += 1

    # Ajustar ancho de columnas según contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # número de columna
        column_letter = get_column_letter(column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'
        return response
    
    
    
    
    
@login_required
def historialAutorizaciones(request):
    usuario = request.user
    historial = AutorizacionSalida.objects.filter(
        autorizador=usuario,
        estado__in=['autorizado', 'rechazado']
    ).select_related('grupo_codigo').order_by('-fecha_autorizacion')

    # Paginación
    paginator = Paginator(historial, 10)  # Mostrar 10 elementos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }
    return render(request, 'Residuos/historialAutorizaciones.html', context)
    
    
    
    
